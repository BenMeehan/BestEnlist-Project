import requests
import validators
import openpyxl

# Function to validate if the entered URL is correct
def validate(url):
    isValid = validators.url(url)

    if isValid == True:
        return True
    else:
        print("Invalid url")
        exit()

# Function to shorten URL
def shorten_url(url):
    query={'url':url}
    response=requests.get("https://benurl.herokuapp.com/shorten",params=query)  # Make GET request to the RESTAPI server

    if response.status_code==500:
        print("\nThere was an error : ",response.text)  # Display any errors
    elif response.status_code==200:
        print("\nShortened URL : ",response.text)   # Display shortened URL if no errors
        return response.text


# Get URL from user
# url=input("Enter the URL to shorten : ")
# validate(url)
# shorten_url(url)

    # OR

# Get URL from EXCEL sheet
xl_path=input("Enter the path of Excel Sheet : ")

# Open Excel Workbook and select sheet
workbook=openpyxl.load_workbook(xl_path)
sheet=workbook.active


# Iterate through the rows and get shortened URL for each entry
for i in range(sheet.max_row):
    value=sheet.cell(row=i+1,column=1).value
    validate(value)
    short_url=shorten_url(value)

    cell=sheet.cell(row=i+1,column=2)
    cell.value=short_url
